#!/usr/bin/env python3
"""Rank catalog products by how often they were ordered on the old B2BWave portal.

Reads the B2BWave order-history export (one row per order line item) and writes a
ranked spreadsheet plus a JSON file the storefront prototype can read.

Usage:  python3 script/rank_products_by_orders.py
"""
import collections
import json
import pathlib

import openpyxl

ROOT = pathlib.Path(__file__).resolve().parent.parent
ORDERS = ROOT / "b2bwave_source_files" / "orders_2026-08-11.xlsx"
CATALOG = ROOT / "db" / "import_data" / "products.json"
XLSX_OUT = ROOT / "docs" / "product_order_frequency.xlsx"
JSON_OUT = ROOT / "db" / "import_data" / "product_order_frequency.json"


def load_line_items():
    wb = openpyxl.load_workbook(ORDERS, read_only=True, data_only=True)
    rows = wb[wb.sheetnames[0]].iter_rows(values_only=True)
    header = [str(h).strip() for h in next(rows)]
    return [dict(zip(header, r)) for r in rows if r[0] is not None]


def s(value):
    return (value or "").strip() if isinstance(value, str) else (value or "")


def main():
    items = load_line_items()
    catalog = {str(p["sku"]).strip(): p for p in json.load(open(CATALOG))}

    stats = collections.defaultdict(
        lambda: {
            "orders": set(),
            "customers": set(),
            "units": 0.0,
            "revenue": 0.0,
            "first": None,
            "last": None,
            "code": "",
        }
    )

    for it in items:
        name = s(it["product_name"])
        if not name:
            continue
        row = stats[name]
        row["code"] = row["code"] or s(it["product_code"])
        row["orders"].add(it["order_id"])
        row["customers"].add(it["customer_id"])
        row["units"] += float(it["product_quantity"] or 0)
        row["revenue"] += float(it["line_total"] or 0)
        date = s(it["order_submitted_date"])
        if date:
            row["first"] = min(row["first"] or date, date)
            row["last"] = max(row["last"] or date, date)

    ranked = sorted(
        stats.items(),
        key=lambda kv: (len(kv[1]["orders"]), kv[1]["units"], kv[1]["revenue"]),
        reverse=True,
    )

    records = []
    for rank, (name, row) in enumerate(ranked, start=1):
        product = catalog.get(row["code"])
        records.append(
            {
                "rank": rank,
                "sku": row["code"],
                "name": name,
                "orders": len(row["orders"]),
                "units": round(row["units"], 2),
                # Named "sales" to match the retail export, so the storefront
                # can read either channel through one code path.
                "sales": round(row["revenue"], 2),
                "customers": len(row["customers"]),
                "first_ordered": row["first"],
                "last_ordered": row["last"],
                "in_catalog": product is not None,
                "category": (product or {}).get("category_path"),
            }
        )

    # --- spreadsheet -------------------------------------------------------
    out = openpyxl.Workbook()
    ws = out.active
    ws.title = "Ranked by orders"
    columns = [
        ("Rank", "rank", 6),
        ("SKU", "sku", 18),
        ("Product", "name", 58),
        ("Times ordered", "orders", 14),
        ("Units sold", "units", 12),
        ("Net sales (USD)", "sales", 16),
        ("Distinct customers", "customers", 18),
        ("First ordered", "first_ordered", 14),
        ("Last ordered", "last_ordered", 14),
        ("In current catalog", "in_catalog", 18),
        ("Category", "category", 34),
    ]
    ws.append([c[0] for c in columns])
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
    ws.freeze_panes = "A2"
    for idx, (label, _key, width) in enumerate(columns, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width

    for rec in records:
        ws.append([
            rec["rank"], rec["sku"], rec["name"], rec["orders"], rec["units"],
            rec["sales"], rec["customers"], rec["first_ordered"], rec["last_ordered"],
            "yes" if rec["in_catalog"] else "no", rec["category"],
        ])
    for row in ws.iter_rows(min_row=2, min_col=6, max_col=6):
        row[0].number_format = '#,##0.00'

    XLSX_OUT.parent.mkdir(parents=True, exist_ok=True)
    out.save(XLSX_OUT)
    total_orders = len({i["order_id"] for i in items})
    all_dates = sorted(d for d in (s(i["order_submitted_date"]) for i in items) if d)
    JSON_OUT.write_text(
        json.dumps(
            {
                "channel": "wholesale",
                "source_file": ORDERS.name,
                "total_orders": total_orders,
                "total_line_items": len(items),
                "first_order_date": all_dates[0],
                "last_order_date": all_dates[-1],
                # Denominators for the storefront percentages. Every product in
                # the export counts, discontinued ones included -- a share of
                # demand should not be inflated by dropping what we stopped
                # selling.
                "total_units": round(sum(r["units"] for r in records), 2),
                "total_sales": round(sum(r["sales"] for r in records), 2),
                "products": records,
            },
            indent=2,
        )
        + "\n"
    )

    print(f"line items      : {len(items)}")
    print(f"distinct orders : {total_orders}")
    print(f"products ranked : {len(records)}")
    print(f"not in catalog  : {sum(1 for r in records if not r['in_catalog'])}")
    print(f"wrote {XLSX_OUT.relative_to(ROOT)}")
    print(f"wrote {JSON_OUT.relative_to(ROOT)}")
    print()
    print(f"{'#':>3}  {'orders':>6}  {'units':>8}  {'revenue':>10}  product")
    for rec in records[:15]:
        print(f"{rec['rank']:>3}  {rec['orders']:>6}  {rec['units']:>8.0f}  "
              f"{rec['sales']:>10,.0f}  {rec['name'][:60]}")


if __name__ == "__main__":
    main()
