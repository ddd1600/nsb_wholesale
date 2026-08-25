#!/usr/bin/env python3
"""Rank products by retail demand, from the Square item exports.

Reads every "retail sales data/items-*.csv" (one row per line item, one file per
year) and writes a ranked spreadsheet plus the JSON the storefront reads.

Retail and wholesale are kept strictly apart. This produces its own file with
its own totals; nothing here is combined with the B2BWave wholesale figures.

Two things the raw export needs before it can be ranked:

  * Non-product rows have to go. "Free Shipping" alone is 1,282 units and would
    otherwise be the top-selling "product" in the company's history.
  * Item names have to be resolved to wholesale SKUs. Between 70% and 94% of
    rows per year carry no SKU at all, and the same product appears under both a
    short point-of-sale name and a long WooCommerce one. The mapping lives in
    db/import_data/retail_item_map.json and is hand-checked, not fuzzy-matched.

Refunds appear as rows with negative quantity and negative sales, so summing
nets them out without special handling.

Usage:  python3 script/rank_retail_sales.py
"""
import collections
import csv
import glob
import json
import pathlib

ROOT = pathlib.Path(__file__).resolve().parent.parent
SOURCE_GLOB = str(ROOT / "retail sales data" / "items-*.csv")
MAP_PATH = ROOT / "db" / "import_data" / "retail_item_map.json"
CATALOG = ROOT / "db" / "import_data" / "products.json"
XLSX_OUT = ROOT / "docs" / "product_retail_frequency.xlsx"
JSON_OUT = ROOT / "db" / "import_data" / "product_retail_frequency.json"


def money(value):
    """'$1,234.56' / '-$40.00' / '($40.00)' -> float."""
    text = (value or "").strip().replace("$", "").replace(",", "")
    if not text:
        return 0.0
    negative = text.startswith("-")
    text = text.lstrip("-")
    if text.startswith("(") and text.endswith(")"):
        text, negative = text[1:-1], True
    try:
        number = float(text)
    except ValueError:
        return 0.0
    return -number if negative else number


def number(value):
    try:
        return float((value or "").strip() or 0)
    except ValueError:
        return 0.0


def main():
    mapping = json.loads(MAP_PATH.read_text())
    excluded = set(mapping["exclude"])
    by_name = mapping["by_name"]
    by_price = mapping["by_name_and_unit_price"]
    catalog = {str(p["sku"]).strip(): p for p in json.load(open(CATALOG))}

    stats = collections.defaultdict(
        lambda: {"units": 0.0, "sales": 0.0, "first": None, "last": None, "names": set()}
    )
    unmapped = collections.defaultdict(lambda: {"units": 0.0, "sales": 0.0})
    totals = {"units": 0.0, "sales": 0.0}
    dates, files, skipped_rows = [], 0, 0

    for path in sorted(glob.glob(SOURCE_GLOB)):
        files += 1
        with open(path, newline="", encoding="utf-8-sig") as handle:
            for row in csv.DictReader(handle):
                name = (row.get("Item") or "").strip()
                if not name or name in excluded:
                    skipped_rows += 1
                    continue

                units = number(row.get("Qty"))
                sales = money(row.get("Net Sales"))

                # Every real product counts toward the denominator, mapped or
                # not: a share of retail demand should not be inflated by
                # quietly dropping the products we no longer carry.
                totals["units"] += units
                totals["sales"] += sales

                date = (row.get("Date") or "").strip()
                if date:
                    dates.append(date)

                sku = resolve(row, name, units, by_name, by_price)
                if sku is None or sku not in catalog:
                    key = name if sku is None else f"{name} [{sku}]"
                    unmapped[key]["units"] += units
                    unmapped[key]["sales"] += sales
                    continue

                entry = stats[sku]
                entry["units"] += units
                entry["sales"] += sales
                entry["names"].add(name)
                if date:
                    entry["first"] = min(entry["first"] or date, date)
                    entry["last"] = max(entry["last"] or date, date)

    records = build_records(stats, catalog, totals)
    write_outputs(records, totals, dates, files, mapping)
    report(records, totals, unmapped, dates, files, skipped_rows)


def resolve(row, name, units, by_name, by_price):
    """Wholesale SKU for a retail row, or None."""
    sku = (row.get("SKU") or "").strip()
    if sku:
        return sku

    if name in by_name:
        return by_name[name]

    prices = by_price.get(name)
    if prices and units:
        unit_price = f"{abs(money(row.get('Gross Sales')) / units):.2f}"
        return prices.get(unit_price)

    return None


def build_records(stats, catalog, totals):
    ranked = sorted(stats.items(), key=lambda kv: (-kv[1]["units"], -kv[1]["sales"]))
    records = []
    for rank, (sku, entry) in enumerate(ranked, start=1):
        product = catalog[sku]
        records.append(
            {
                "rank": rank,
                "sku": sku,
                "name": product["name"],
                "units": round(entry["units"], 2),
                "sales": round(entry["sales"], 2),
                "units_share": entry["units"] / totals["units"] if totals["units"] else 0.0,
                "sales_share": entry["sales"] / totals["sales"] if totals["sales"] else 0.0,
                "first_sold": entry["first"],
                "last_sold": entry["last"],
                "source_names": sorted(entry["names"]),
                "category": product.get("category_path"),
            }
        )
    return records


def write_outputs(records, totals, dates, files, mapping):
    import openpyxl

    JSON_OUT.write_text(
        json.dumps(
            {
                "channel": "retail",
                "source": "Square item exports, retail sales data/items-*.csv",
                "source_files": files,
                "first_sale_date": min(dates),
                "last_sale_date": max(dates),
                "total_units": round(totals["units"], 2),
                "total_sales": round(totals["sales"], 2),
                "products": records,
            },
            indent=2,
        )
        + "\n"
    )

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Retail by units"
    columns = [
        ("Rank", "rank", 6), ("SKU", "sku", 24), ("Product", "name", 58),
        ("Units sold", "units", 12), ("% of units", "units_share", 12),
        ("Net sales (USD)", "sales", 16), ("% of sales", "sales_share", 12),
        ("First sold", "first_sold", 13), ("Last sold", "last_sold", 13),
        ("Category", "category", 30),
    ]
    sheet.append([c[0] for c in columns])
    for cell in sheet[1]:
        cell.font = openpyxl.styles.Font(bold=True)
    sheet.freeze_panes = "A2"
    for index, (_label, _key, width) in enumerate(columns, start=1):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(index)].width = width

    for record in records:
        sheet.append([record[key] for _label, key, _width in columns])
    for row in sheet.iter_rows(min_row=2):
        row[4].number_format = "0.0%"
        row[6].number_format = "0.0%"
        row[5].number_format = "#,##0.00"

    XLSX_OUT.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(XLSX_OUT)


def report(records, totals, unmapped, dates, files, skipped_rows):
    mapped_units = sum(r["units"] for r in records)
    mapped_sales = sum(r["sales"] for r in records)
    print(f"source files        : {files}  ({min(dates)} -> {max(dates)})")
    print(f"non-product rows    : {skipped_rows} skipped (shipping, fees, tests)")
    print(f"retail totals       : {totals['units']:,.0f} units / ${totals['sales']:,.0f}")
    print(f"matched to catalog  : {len(records)} products, "
          f"{mapped_units / totals['units'] * 100:.1f}% of units, "
          f"{mapped_sales / totals['sales'] * 100:.1f}% of sales")
    print(f"wrote {JSON_OUT.relative_to(ROOT)}")
    print(f"wrote {XLSX_OUT.relative_to(ROOT)}")
    print()
    print(f"{'#':>3} {'units':>7} {'%u':>6} {'sales':>10} {'%$':>6}  product")
    for record in records[:15]:
        print(f"{record['rank']:>3} {record['units']:>7,.0f} {record['units_share'] * 100:>5.1f}% "
              f"{record['sales']:>10,.0f} {record['sales_share'] * 100:>5.1f}%  {record['name'][:46]}")

    if unmapped:
        print()
        print("UNMAPPED (real retail sales with no product in the wholesale catalog):")
        for name, totals_row in sorted(unmapped.items(), key=lambda kv: -kv[1]["sales"])[:20]:
            print(f"  {totals_row['units']:>7,.0f}u ${totals_row['sales']:>9,.0f}  {name[:64]}")


if __name__ == "__main__":
    main()
